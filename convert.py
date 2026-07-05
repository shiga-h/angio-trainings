import openpyxl
import json
import re
import sys
import datetime

TARGET_SHEET_NAME = "研究会一覧"
MONTH_PATTERN = re.compile(r"^(1[0-2]|[1-9])月$")
SRC = sys.argv[1] if len(sys.argv) > 1 else "2026年度研修会_統合版.xlsx"
OUT = "data.json"


# ============================================================
# レイアウト判定
# ============================================================

INTEGRATED_KEYS = ("ID", "開始日", "名称")   # これらが揃っていれば統合版
MONTHLY_KEYS    = ("日時", "名称")            # これらが揃っていれば月別


def detect_layout(header_row):
    """1行目のヘッダーからレイアウトを判定。'integrated' / 'monthly' / None"""
    hs = {str(x).strip() for x in header_row if x is not None and str(x).strip()}
    if all(k in hs for k in INTEGRATED_KEYS):
        return "integrated"
    if all(k in hs for k in MONTHLY_KEYS):
        return "monthly"
    return None


# ============================================================
# 年度推定（月別シートで年が省略されている場合の推定用）
# ============================================================

def current_fiscal_start_year() -> int:
    """今日の日付から現在の年度開始年（4月始まりの前提）を推定。"""
    today = datetime.date.today()
    return today.year if today.month >= 4 else today.year - 1


def sheet_base_year(sheet_name: str, fiscal_start: int) -> int:
    """シート名 '4月'〜'3月' からそのシートの日付の基準年を返す。"""
    m = MONTH_PATTERN.match(sheet_name)
    if not m:
        return fiscal_start
    month = int(m.group(1))
    # 4-12月は年度開始年、1-3月は翌年
    return fiscal_start if 4 <= month <= 12 else fiscal_start + 1


# ============================================================
# 日時文字列パーサー（月別シート用）
# ============================================================

_TIME_RE = re.compile(r"\d{1,2}:\d{2}(?:\s*[〜~ー\-–]\s*\d{1,2}:\d{2})?")
_DOW_RE  = re.compile(r"[（(][月火水木金土日祝][）)]")
_DATE_RE = re.compile(r"(?:(\d{4})[/./\-年])?(\d{1,2})[/./\-月](\d{1,2})日?")


def parse_date_field(v, base_year: int) -> tuple[str, str, str]:
    """日時セル値から (dateStart, dateEnd, time) を返す。

    - Date/datetime オブジェクトはそのまま日付として使う
    - 文字列は正規表現で時刻と日付を切り出す
    - 日付が2つ見つかれば1つめ=start, 2つめ=end
    - 年が省略されていれば base_year を採用
    - 日付が全く抽出できなかった場合は生文字列を time に残す（表示上見えるように）
    """
    if v is None:
        return "", "", ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d"), "", ""
    s = str(v).strip()
    if not s:
        return "", "", ""
    original = s

    # 1. 時刻抽出
    time_match = _TIME_RE.search(s)
    time_str = time_match.group(0) if time_match else ""
    if time_match:
        s = s[:time_match.start()] + s[time_match.end():]

    # 2. 曜日マーカー除去
    s = _DOW_RE.sub("", s)

    # 3. 日付抽出
    matches = _DATE_RE.findall(s)

    def to_date(match) -> str:
        year_str, month, day = match
        year = int(year_str) if year_str else base_year
        return f"{year:04d}-{int(month):02d}-{int(day):02d}"

    date_start = to_date(matches[0]) if matches else ""
    date_end   = to_date(matches[1]) if len(matches) >= 2 else ""

    # フォールバック: 日付が全く取れなかったら、生文字列を time に置いておく
    if not date_start:
        time_str = original

    return date_start, date_end, time_str


def clean_url(v) -> str:
    """URL 値を検証。http:// または https:// で始まらないものは空扱い。"""
    if v is None:
        return ""
    s = str(v).strip()
    if s.startswith(("http://", "https://")):
        return s
    return ""


# ============================================================
# 行パーサー
# ============================================================

def fmt_date(d) -> str:
    if not d:
        return ""
    if hasattr(d, "strftime"):
        return d.strftime("%Y-%m-%d")
    return str(d)


def parse_integrated_row(row):
    """統合版レイアウト (13列)"""
    if len(row) < 13:
        row = tuple(row) + (None,) * (13 - len(row))
    (rid, date_start, date_end, time_, name, theme, loc, fmt,
     fee, points, url, tag1, tag2) = row[:13]
    if not name:
        return None
    tags = [t for t in [tag1, tag2] if t]
    return {
        "id": rid,
        "dateStart": fmt_date(date_start),
        "dateEnd": fmt_date(date_end),
        "time": time_ or "",
        "name": name,
        "theme": theme or "",
        "loc": loc or "",
        "format": fmt or "未定",
        "fee": fee or "",
        "points": points or "",
        "url": clean_url(url),
        "tags": tags,
    }


def parse_monthly_row(row, base_year: int):
    """月別レイアウト (8列): [日時, 名称, テーマ, 開催場所, 開催形式, 参加費, 認定単位, HP]"""
    if len(row) < 8:
        row = tuple(row) + (None,) * (8 - len(row))
    (date_raw, name, theme, loc, fmt, fee, points, url) = row[:8]
    if not name:
        return None
    date_start, date_end, time_ = parse_date_field(date_raw, base_year)
    return {
        "id": "",
        "dateStart": date_start,
        "dateEnd": date_end,
        "time": time_,
        "name": name,
        "theme": theme or "",
        "loc": loc or "",
        "format": fmt or "未定",
        "fee": fee or "",
        "points": points or "",
        "url": clean_url(url),
        "tags": [],
    }


# ============================================================
# シート → 行リスト
# ============================================================

def read_sheet(ws, sheet_name: str, fiscal_start: int):
    """指定シートを読んで dict のリストを返す。レイアウトはヘッダーから自動判別。"""
    # ヘッダー抽出
    header_row = ()
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        header_row = row
        break
    header_preview = [str(x) if x is not None else "" for x in header_row]
    layout = detect_layout(header_row)
    print(f"    ヘッダー: {header_preview}  → レイアウト: {layout}")

    if layout is None:
        return []

    max_col = 13 if layout == "integrated" else 8
    base_year = sheet_base_year(sheet_name, fiscal_start)

    out = []
    sample_logged = False
    for row in ws.iter_rows(min_row=2, max_col=max_col, values_only=True):
        if layout == "integrated":
            item = parse_integrated_row(row)
        else:
            item = parse_monthly_row(row, base_year)
        if item is None:
            continue
        # 各シートの最初の1件だけパース結果をログ出力
        if not sample_logged:
            print(f"    サンプル raw 日時: {row[0]!r}")
            print(f"    サンプル パース後 : dateStart={item['dateStart']!r}, "
                  f"dateEnd={item['dateEnd']!r}, time={item['time']!r}")
            sample_logged = True
        out.append(item)
    return out


# ============================================================
# ワークブック → 行リスト
# ============================================================

def collect_rows(wb):
    """優先順位でデータを収集する。

    1. '研究会一覧' シートがあればそれを単体で使う
    2. '4月'〜'3月' が2つ以上あれば全部読んで結合
    3. シートが1つだけならそれを使う
    4. それ以外はエラー
    """
    fiscal_start = current_fiscal_start_year()
    print(f"年度開始年（推定）: {fiscal_start}")

    names = wb.sheetnames

    if TARGET_SHEET_NAME in names:
        print(f"シート '{TARGET_SHEET_NAME}' を使用します")
        return read_sheet(wb[TARGET_SHEET_NAME], TARGET_SHEET_NAME, fiscal_start)

    monthly = [n for n in names if MONTH_PATTERN.match(n)]
    if len(monthly) >= 2:
        print(f"月別シート {len(monthly)}枚 ({monthly}) を統合します")
        all_rows = []
        for n in monthly:
            sheet_rows = read_sheet(wb[n], n, fiscal_start)
            print(f"  {n}: {len(sheet_rows)}件")
            all_rows.extend(sheet_rows)
        return all_rows

    if len(names) == 1:
        chosen = names[0]
        print(f"'{TARGET_SHEET_NAME}' なし → 唯一のシート '{chosen}' を使用します")
        return read_sheet(wb[chosen], chosen, fiscal_start)

    raise SystemExit(
        f"ERROR: '{TARGET_SHEET_NAME}' シートも月別シート群 (2つ以上の '〇月') もなく、"
        f"シートが複数あって特定できません。\n"
        f"  ファイル内のシート: {names}\n"
        f"  対処: 該当シートを '{TARGET_SHEET_NAME}' にリネームするか、"
        f"月名 '4月' 等のシートで揃えてください。"
    )


# ============================================================
# main
# ============================================================

wb = openpyxl.load_workbook(SRC, data_only=True)
rows = collect_rows(wb)

with open(OUT, "w", encoding="utf-8") as f:
    json.dump(rows, f, ensure_ascii=False, indent=2)

print(f"{len(rows)}件を{OUT}に書き出しました")
